# 2.6 ESTUDO DEMANDA
from janelas.janela_estudo_pedidos import *
from PyQt5.QtWidgets import QMainWindow, QFileDialog, QMessageBox, QWidget, QShortcut
from PyQt5.QtCore import QDate
from banco_de_dados_workflow import BDWorkflow
from banco_de_dados_oracle import BDBohm
import openpyxl
import xlsxwriter
import json
import datetime
from openpyxl import load_workbook


path_download = ''
with open('setup.json', 'r') as file:
    d1_json = file.read()
    d1_json = json.loads(d1_json)

for x, y in d1_json.items():
    path_download = y['relatorio']


class EstudoPedido(QMainWindow, Ui_estudo_pedido):
    def __init__(self, widget_relatorios, parent=None):
        super().__init__(parent)
        super().setupUi(self)
        self.btnCalcular.clicked.connect(self.analisar)

    def analisar(self):
        try:
            pedido = self.inputPedido.text()
            if len(pedido) == 5:
                self.calcular(pedido)
            else:
                QMessageBox.about(self, "Erro", "Digitar um pedido")
        except:
            QMessageBox.about(self, "Erro", "Erro B.D.")

    def calcular(self, pedido):
        itens = []
        itens_ops = []
        itens_estrutura = []
        # PEGA ITENS DO PEDIDO
        respostas = BDBohm().pegar_itens_pedido(pedido)
        indice_item = 1
        for resposta in respostas:
            item = {}
            if len(str(indice_item)) == 1:
                item['indice'] = f"0{str(indice_item)}"
            else:
                item['indice'] = str(indice_item)
            item['pai'] = ''
            item['preco_venda'] = resposta[2]
            item['produto'] = resposta[0]
            item['quantidade'] = resposta[1]
            item['correto'] = False
            itens.append(item)
            indice_item += 1

        # Procurar sub item
        index = 1
        while index > 0:
            index = 0
            for item in itens:
                produto = item['produto']
                indice = item['indice']
                quantidade_produto = item['quantidade']
                resposta1 = BDBohm().procurar_chave_processo(produto)
                if not resposta1:
                    item['tipo'] = 0
                    item['processo'] = ''
                    itens_estrutura.append(item)
                else:
                    item['tipo'] = 1
                    chave_processo = resposta1[0]
                    item['processo'] = chave_processo
                    item['correto'] = True
                    itens_estrutura.append(item)
                    index = 1
                    respostas2 = BDBohm().procurar_subitens(chave_processo)
                    indice_sub_item = 1
                    for resposta2 in respostas2:
                        item = {}
                        if len(str(indice_sub_item)) == 1:
                            item['indice'] = f"{indice}.0{str(indice_sub_item)}"
                        else:
                            item['indice'] = f"{indice}.{str(indice_sub_item)}"
                        item['pai'] = f"{produto}"
                        item['preco_venda'] = ''
                        item['produto'] = resposta2[0]
                        item['quantidade'] = resposta2[1] * quantidade_produto
                        itens_ops.append(item)
                        indice_sub_item += 1
            itens = itens_ops
            itens_ops = []

        # PEGA CUSTOS DE PROCESSOS OU ULTIMA OC
        index = 0
        respostas = BDBohm().pegar_custos_setores()
        custos_setores = []
        for resposta in respostas:
            dict = {}
            dict['chave'] = resposta[0]
            dict['setor'] = resposta[1]
            dict['custo'] = resposta[2]
            custos_setores.append(dict)

        for item in itens_estrutura:
            if item['tipo'] == 1:
                processo = item['processo']
                resultados = BDBohm().pegar_custos_processo(processo)
                custo_setor = 0
                custo = 0
                for resultado in resultados:
                    setor = resultado[0]
                    setup = resultado[1]
                    ciclo = resultado[2]
                    for n in custos_setores:
                        if n['chave'] == setor:
                            custo_setor = n['custo']
                    custo = custo + (custo_setor * (ciclo/60)) + (custo_setor * (setup/60))
                item['custo'] = custo
            else:
                produto = item['produto']
                respostas = BDBohm().pegar_ultimo_custo_de_compra(produto)
                for resposta in respostas:
                    cubagem = resposta[4]
                    unidade_producao = resposta[3]
                    unidade_compra = resposta[5]
                    moeda = resposta[6]
                    valor_moeda = 1
                    if moeda != '0':
                        respostas2 = BDBohm().cotacao_moeda(moeda)
                        for resposta2 in respostas2:
                            valor_moeda = resposta2[0]
                    item['correto'] = True
                    if unidade_compra == unidade_producao:
                        item['custo'] = resposta[1] * valor_moeda
                    else:
                        item['custo'] = resposta[1] * valor_moeda * cubagem
            index += 1
        lista_ordenada = sorted(itens_estrutura, key=lambda d: d['indice'])
        gerar_relatorio = True
        for item in lista_ordenada:
            verificar = item["correto"]
            produto_erro = item["produto"]
            if not verificar:
                gerar_relatorio = False
                QMessageBox.about(self, "Erro", f"Item {produto_erro} não tem custo cadastrado")
                break
        if gerar_relatorio:
            self.gerar_relatorio(path_download, lista_ordenada)

    def gerar_relatorio(self, caminho, lista):
        fazer = True
        indice = 1
        indice_text = ''
        workbook = xlsxwriter.Workbook(caminho + 'custo_pedido.xlsx')
        while fazer:
            fazer = False
            if len(str(indice)) == 1:
                indice_text = f"0{str(indice)}"
            elif len(str(indice)) == 2:
                indice_text = f"{str(indice)}"
            else:
                indice_text = 'erro'
                QMessageBox.about(self, "Erro", "Pedido contem mais de 100 itens o programa não suporta")
            worksheet = workbook.add_worksheet(f'Item {indice_text}')
            worksheet.write(0, 0, 'Valor de Venda')
            worksheet.write(0, 1, 'Custo Total')
            worksheet.write(0, 2, 'Custo/Venda')
            worksheet.write(0, 1, 'indice')
            worksheet.write(0, 4, 'indice')
            worksheet.write(0, 5, 'pai')
            worksheet.write(0, 6, 'preco_venda')
            worksheet.write(0, 7, 'custo')
            worksheet.write(0, 8, 'produto')
            worksheet.write(0, 9, 'quantidade')
            worksheet.write(0, 10, 'tipo')
            worksheet.write(0, 11, 'processo')
            worksheet.write(0, 12, 'soma custos')
            row = 1
            try:
                soma_custo = 0
                soma_preco = 0
                for item in lista:
                    indice_item = item['indice']
                    if indice_item[:2] == indice_text:
                        fazer = True
                        if item['preco_venda']:
                            preco_venda = item['preco_venda']
                        else:
                            preco_venda = 0
                        custo = item['custo']
                        quantidade = item['quantidade']
                        worksheet.write(row, 4, item['indice'])
                        worksheet.write(row, 5, item['pai'])
                        worksheet.write(row, 6, preco_venda)
                        worksheet.write(row, 7, custo)
                        worksheet.write(row, 8, item['produto'])
                        worksheet.write(row, 9, quantidade)
                        worksheet.write(row, 10, item['tipo'])
                        worksheet.write(row, 11, item['processo'])
                        worksheet.write(row, 12, custo * quantidade)
                        soma_preco = soma_preco + (preco_venda * quantidade)
                        soma_custo = soma_custo + (custo * quantidade)
                        row += 1
                if fazer:
                    worksheet.write(1, 0, soma_preco)
                    worksheet.write(1, 1, soma_custo)
                    worksheet.write(1, 2, (soma_custo/soma_preco))
                indice += 1
            except:
                QMessageBox.about(self, "Erro", "Erro ao puchar os dados")
        workbook.close()

        wb = load_workbook(caminho + 'custo_pedido.xlsx')
        if f'Item {indice_text}' in wb.sheetnames:
            wb.remove(wb[f'Item {indice_text}'])
        wb.save(caminho + 'custo_pedido.xlsx')

        QMessageBox.about(self, "Sucesso", "Arquivo Gerado")
