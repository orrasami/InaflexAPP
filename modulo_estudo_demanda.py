# 2.6 ESTUDO DEMANDA
from janelas.janela_estudo_demanda import *
from PyQt5.QtWidgets import QMainWindow, QFileDialog, QMessageBox, QWidget, QShortcut
from PyQt5.QtCore import QDate
from banco_de_dados_workflow import BDWorkflow
from banco_de_dados_oracle import BDBohm
import openpyxl
import xlsxwriter
import json
import datetime


path_download = ''
with open('setup.json', 'r') as file:
    d1_json = file.read()
    d1_json = json.loads(d1_json)

for x, y in d1_json.items():
    path_download = y['relatorio']


class EstudoDemanda(QMainWindow, Ui_estudo_demanda):
    def __init__(self, widget_relatorios, parent=None):
        super().__init__(parent)
        super().setupUi(self)
        self.btnSelect.clicked.connect(self.abrir_arquivo)
        hoje = datetime.datetime.now()
        data_padrao = hoje - datetime.timedelta(days=365)
        data = QDate(data_padrao.year, data_padrao.month, data_padrao.day)
        self.data_inicial.setDate(data)

    def abrir_arquivo(self):
        try:
            arquivo = QFileDialog.getOpenFileName(
                self.centralWidget(),
                'Escolher Arquivo',
                path_download,
            )
            caminho = arquivo[0]
            wb = openpyxl.load_workbook(caminho)
            ws = wb.worksheets[0]
            row_count = ws.max_row
            lista = []
            for i in range(row_count):
                valor = self.pegar_valores(ws, i + 1)
                lista.append(valor)
            self.gerar_relatorio(path_download, lista)
        except:
            QMessageBox.about(self, "Erro", "Nenhum arquivo selecionado")

    @staticmethod
    def pegar_valores(ws, linha):
        valor = ws.cell(row=linha, column=1).value
        return valor

    def gerar_relatorio(self, caminho, lista):
        data_qdate = self.data_inicial.date()
        data_py = data_qdate.toPyDate()
        data = data_py.strftime("%d/%m/%Y")
        workbook = xlsxwriter.Workbook(caminho + 'demanda.xlsx')
        worksheet = workbook.add_worksheet('Relatorio')
        worksheet.write(0, 0, 'item')
        worksheet.write(0, 1, 'descricao')
        worksheet.write(0, 2, 'qtd orcados')
        worksheet.write(0, 3, 'qtd orcamentos')
        worksheet.write(0, 4, 'qtd vendidos')
        worksheet.write(0, 5, 'qtd pedidos')
        worksheet.write(0, 6, 'estoque atual')
        worksheet.write(0, 7, 'estoque minimo')
        worksheet.write(0, 8, 'estoque na data inicial')
        worksheet.write(0, 9, 'entrada de compras')
        worksheet.write(0, 10, 'entrada de importação')
        worksheet.write(0, 11, 'entrada de ops')
        worksheet.write(0, 12, 'demanda')
        worksheet.write(0, 13, 'custo')
        row = 1
        try:
            for item in lista:
                (estoque_data_inicial, entrada_compras, entrada_importacao, entrada_ops, qtd_orcados,
                 qtd_orcamentos, qtd_vendidos, qtd_pedidos, estoque_atual, estoque_minimo, demanda,
                 descricao, custo) = BDBohm().pegar_valores_demanda(item, data)
                worksheet.write(row, 0, item)
                worksheet.write(row, 1, descricao)
                worksheet.write(row, 2, qtd_orcados)
                worksheet.write(row, 3, qtd_orcamentos)
                worksheet.write(row, 4, qtd_vendidos)
                worksheet.write(row, 5, qtd_pedidos)
                worksheet.write(row, 6, estoque_atual)
                worksheet.write(row, 7, estoque_minimo)
                worksheet.write(row, 8, estoque_data_inicial)
                worksheet.write(row, 9, entrada_compras)
                worksheet.write(row, 10, entrada_importacao)
                worksheet.write(row, 11, entrada_ops)
                worksheet.write(row, 12, demanda)
                worksheet.write(row, 13, custo)
                row += 1
            row += 1
            worksheet.write(row, 0, f"Data inicial {data}")
            row += 1
            worksheet.write(row, 0, "Demanda = Quantidade na Data Inicial + Entrada Compras + Entrada OP + "
                                    "Entrada Importação - Estoque Atual")
            workbook.close()
            QMessageBox.about(self, "Sucesso", "Arquivo Gerado")
        except:
            QMessageBox.about(self, "Erro", "Erro ao puchar os dados")
